"""Excel export — multi-sheet close package for download."""

import io

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment, numbers

from agent.tools import (
    COST_CATEGORIES,
    calculate_accruals,
    calculate_net_down,
    calculate_outlook,
    generate_outlook_load_file,
    get_exceptions,
)


HEADER_FILL = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DOLLAR_FMT = '#,##0'


def _style_header(ws):
    """Apply green header style to the first row of a worksheet."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")


def _auto_width(ws):
    """Auto-fit column widths (approximate)."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 30)


def _apply_dollar_format(ws, dollar_cols):
    """Apply dollar number format to specified columns (1-indexed)."""
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in dollar_cols:
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = DOLLAR_FMT


def generate_close_package(business_unit: str = "all") -> bytes:
    """Generate the full close package as an Excel workbook (bytes).

    Sheets:
    1. Accrual Summary — per-well gross/net accruals
    2. Net-Down Report — WI% adjustments
    3. Outlook Summary — future outlook per well
    4. OneStream Load   — monthly grid
    5. Exception Report
    """
    output = io.BytesIO()

    accrual_result = calculate_accruals(business_unit)
    net_down_result = calculate_net_down(business_unit)
    outlook_result = calculate_outlook(business_unit)
    load_result = generate_outlook_load_file(business_unit)
    exception_result = get_exceptions(business_unit)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # --- Sheet 1: Accrual Summary ---
        accrual_rows = []
        for r in accrual_result["accruals"]:
            accrual_rows.append({
                "WBS Element": r["wbs_element"],
                "Well Name": r["well_name"],
                "Business Unit": r["business_unit"],
                "WI%": r["wi_pct"],
                "Drill Gross": r["drill_gross_accrual"],
                "Comp Gross": r["comp_gross_accrual"],
                "FB Gross": r["fb_gross_accrual"],
                "HU Gross": r["hu_gross_accrual"],
                "Total Gross": r["total_gross_accrual"],
                "Total Net": r["total_net_accrual"],
                "Prior Gross": r["prior_gross_accrual"],
            })
        df_accruals = pd.DataFrame(accrual_rows)
        df_accruals.to_excel(writer, sheet_name="Accrual Summary", index=False)
        ws = writer.sheets["Accrual Summary"]
        _style_header(ws)
        _apply_dollar_format(ws, [5, 6, 7, 8, 9, 10, 11])
        _auto_width(ws)

        # --- Sheet 2: Net-Down Report ---
        if net_down_result["adjustments"]:
            nd_rows = []
            for a in net_down_result["adjustments"]:
                nd_rows.append({
                    "WBS Element": a["wbs_element"],
                    "Well Name": a["well_name"],
                    "Total System Cost": a["total_system_cost"],
                    "System WI%": a["system_wi_pct"],
                    "Actual WI%": a["actual_wi_pct"],
                    "WI Discrepancy": a["wi_discrepancy"],
                    "Net-Down Adjustment": a["net_down_adjustment"],
                    "Adjusted Net Cost": a["adjusted_net_cost"],
                })
            df_nd = pd.DataFrame(nd_rows)
        else:
            df_nd = pd.DataFrame({"Note": ["No WI% mismatches found"]})
        df_nd.to_excel(writer, sheet_name="Net-Down Report", index=False)
        ws = writer.sheets["Net-Down Report"]
        _style_header(ws)
        if net_down_result["adjustments"]:
            _apply_dollar_format(ws, [3, 7, 8])
        _auto_width(ws)

        # --- Sheet 3: Outlook Summary ---
        outlook_rows = []
        for r in outlook_result["outlook"]:
            outlook_rows.append({
                "WBS Element": r["wbs_element"],
                "Well Name": r["well_name"],
                "Business Unit": r["business_unit"],
                "WI%": r["wi_pct"],
                "Total Ops Budget": r["total_ops_budget"],
                "Total Future Outlook": r["total_future_outlook"],
            })
        df_outlook = pd.DataFrame(outlook_rows)
        df_outlook.to_excel(writer, sheet_name="Outlook Summary", index=False)
        ws = writer.sheets["Outlook Summary"]
        _style_header(ws)
        _apply_dollar_format(ws, [5, 6])
        _auto_width(ws)

        # --- Sheet 4: OneStream Load ---
        df_load = load_result["load_file"]
        df_load.to_excel(writer, sheet_name="OneStream Load", index=False)
        ws = writer.sheets["OneStream Load"]
        _style_header(ws)
        _auto_width(ws)

        # --- Sheet 5: Exception Report ---
        if exception_result["exceptions"]:
            df_exc = pd.DataFrame(exception_result["exceptions"])
            col_order = ["wbs_element", "well_name", "exception_type", "severity", "detail"]
            df_exc = df_exc[[c for c in col_order if c in df_exc.columns]]
            df_exc.columns = ["WBS Element", "Well Name", "Type", "Severity", "Detail"]
        else:
            df_exc = pd.DataFrame({"Note": ["No exceptions"]})
        df_exc.to_excel(writer, sheet_name="Exception Report", index=False)
        ws = writer.sheets["Exception Report"]
        _style_header(ws)
        _auto_width(ws)

    return output.getvalue()
